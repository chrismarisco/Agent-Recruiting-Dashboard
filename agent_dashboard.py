import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import re
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Configure page
st.set_page_config(
    page_title="RealtyMetric Solutions - Agent Recruiting Dashboard",
    page_icon="🏠",
    layout="wide"
)

#---------------------------
# Custom Branding & Styling
#---------------------------
st.markdown("""
<style>
    .main {
        padding-top: 10px;
    }
    .branding-header {
        background: linear-gradient(135deg, #1B2A4A, #2E5090);
        color: white;
        padding: 20px 30px;
        border-radius: 10px;
        margin-bottom: 20px;
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    .branding-header h1 {
        margin: 0;
        font-size: 28px;
        color: white;
    }
    .branding-header p {
        margin: 5px 0 0 0;
        color: #A8C4E0;
        font-size: 14px;
    }
    .footer {
        margin-top: 40px;
        padding: 15px 30px;
        border-top: 1px solid #E0E0E0;
        color: #888;
        font-size: 12px;
        text-align: center;
    }
    .stButton > button {
        background-color: #2E5090;
        color: white;
    }
    .stButton > button:hover {
        background-color: #1B2A4A;
    }
</style>
""", unsafe_allow_html=True)

#---------------------------
# Helper Functions
#---------------------------
def dollar_to_numeric(x):
    """Convert dollar strings to numeric values"""
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        return float(re.sub(r'[$,]', '', x))
    return float(x)

def clean_agent_data(raw_df):
    """Clean and prepare agent data"""
    df = raw_df.copy()
    for col in ['TotalVolume', 'ListSideVolume', 'SellSideVolume']:
        if col in df.columns:
            df[col] = df[col].apply(dollar_to_numeric)
    for col in ['TotalCount', 'DaysOnMarket']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    if 'MovedDate' in df.columns:
        df['MovedDate'] = pd.to_datetime(df['MovedDate'], format='%m/%d/%Y', errors='coerce')
        df['RecencyDays'] = (pd.Timestamp.now() - df['MovedDate']).dt.days
    df['FirstName'] = df['FirstName'].fillna('Unnamed').replace('', 'Unnamed')
    df['LastName'] = df['LastName'].fillna('Agent').replace('', 'Agent')
    if 'City' in df.columns:
        df['City_clean'] = df['City'].str.strip().str.title()
    # Remove duplicate agents (keep highest TotalVolume)
    if 'TotalVolume' in df.columns:
        df = df.sort_values('TotalVolume', ascending=False)
    df = df.drop_duplicates(subset=['FirstName', 'LastName', 'OfficeName'], keep='first')
    return df

def safe_rescale(series):
    """Safely rescale a series to 0-1 range"""
    if series.empty or series.isna().all():
        return pd.Series([0.5] * len(series), index=series.index)
    min_val, max_val = series.min(), series.max()
    if pd.isna(min_val) or pd.isna(max_val) or min_val == max_val:
        return pd.Series([0.5] * len(series), index=series.index)
    return (series - min_val) / (max_val - min_val)

def format_currency(x):
    """Format currency values"""
    if pd.isna(x): return "—"
    return f"${x:,.0f}"

def format_integer(x):
    """Format integer values"""
    if pd.isna(x): return "—"
    return f"{int(x):,}"

def has_county_data(df):
    """Check if real county data exists (not all Unknown)"""
    counties = df['County'].dropna().unique()
    return not (len(counties) == 0 or (len(counties) == 1 and counties[0] == 'Unknown'))

@st.cache_data
def load_and_prepare_data():
    """Load and prepare all data"""
    try:
        raw_data = pd.read_csv("CA Dashboard.csv")
        agents = clean_agent_data(raw_data)
        try:
            city2county = pd.read_csv("CA_City_to_County_Mapping.csv")
            city2county['City_clean'] = city2county['City'].str.strip().str.title()
            city2county['County'] = city2county['County'].str.strip().str.title()
            city2county = city2county[['City_clean', 'County']].drop_duplicates('City_clean')
        except FileNotFoundError:
            st.warning("City to County mapping file not found. All cities will be marked as 'Unknown'.")
            city2county = pd.DataFrame(columns=['City_clean', 'County'])
        if 'City_clean' in agents.columns:
            unmapped = agents[['City_clean']].drop_duplicates()
            unmapped = unmapped[~unmapped['City_clean'].isin(city2county['City_clean'])]
            unmapped['County'] = 'Unknown'
            city2county_full = pd.concat([city2county, unmapped], ignore_index=True).drop_duplicates('City_clean')
            agents = agents.merge(city2county_full, on='City_clean', how='left')
            agents['County'] = agents['County'].fillna('Unknown').replace('', 'Unknown')
        else:
            agents['County'] = 'Unknown'
            city2county_full = pd.DataFrame(columns=['City_clean', 'County'])
        return agents, city2county_full
    except FileNotFoundError:
        st.error("Please upload 'CA Dashboard.csv' file to proceed.")
        return None, None

#---------------------------
# Authentication
#---------------------------
def check_password():
    """Returns `True` if the user had the correct password."""
    
    # Initialize password_correct in session state if not present
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    
    # If already authenticated, stay authenticated
    if st.session_state["password_correct"]:
        return True
    
    # Show password input
    password_input = st.text_input("Password", type="password", key="password_input")
    st.write("*Please contact your administrator for access.*")
    
    # Check password when user enters it
    if password_input:
        if password_input == st.secrets["password"]:
            st.session_state["password_correct"] = True
            st.rerun()
        else:
            st.error("😕 Password incorrect")
            return False
    
    return False

#---------------------------
# Excel Export
#---------------------------
def export_to_excel(df, include_summary=False, top_n=None):
    """Create a professionally formatted Excel export"""
    output = BytesIO()
    
    if top_n:
        df = df.head(top_n)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ranked Agents', startrow=2 if include_summary else 1)
        ws = writer.sheets['Ranked Agents']
        
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        subtitle_font = Font(size=11, color="666666", italic=True)
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        alt_row_fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
        
        ws['A1'] = 'RealtyMetric Solutions – Agent Recruiting Report'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws.row_dimensions[1].height = 30
        
        start_row = 3 if include_summary else 2
        if include_summary:
            ws['A2'] = f'Executive Summary - Top {top_n if top_n else len(df)} Recruiting Targets | Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
            ws['A2'].font = subtitle_font
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
            ws.row_dimensions[2].height = 20
        
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[start_row].height = 25
        
        for i, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() or 0)
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 30)
        
        if 'TotalVolume' in df.columns:
            tv_col = list(df.columns).index('TotalVolume') + 1
            for row in range(start_row + 1, len(df) + start_row + 1):
                ws.cell(row=row, column=tv_col).number_format = '$#,##0'
        
        if 'Final_Score' in df.columns:
            sc_col = list(df.columns).index('Final_Score') + 1
            for row in range(start_row + 1, len(df) + start_row + 1):
                ws.cell(row=row, column=sc_col).number_format = '0.0000'
        
        for row in range(start_row + 1, len(df) + start_row + 1, 2):
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = alt_row_fill
    
    output.seek(0)
    return output

#---------------------------
# Email Report Function
#---------------------------
def send_email_report(recipient_email, attachment_data, attachment_filename, attachment_type="excel"):
    """Send report via email (requires email configuration in secrets)"""
    try:
        # Get email config from secrets
        smtp_server = st.secrets.get("smtp_server", "smtp.gmail.com")
        smtp_port = st.secrets.get("smtp_port", 587)
        sender_email = st.secrets.get("sender_email", "")
        sender_password = st.secrets.get("sender_password", "")
        
        if not sender_email or not sender_password:
            return False, "Email configuration not found. Please contact support."
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = f"RealtyMetric Solutions - Agent Recruiting Report - {date.today().strftime('%B %d, %Y')}"
        
        body = f"""
Dear Valued Client,

Please find attached your RealtyMetric Solutions Agent Recruiting Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}.

This report contains your customized agent rankings based on your specified criteria.

If you have any questions or need assistance, please don't hesitate to contact us at support@realtymetricsolutions.com.

Best regards,
The RealtyMetric Solutions Team

---
© 2026 RealtyMetric Solutions | Confidential
"""
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach file
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={attachment_filename}')
        msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        
        return True, "Email sent successfully!"
        
    except Exception as e:
        return False, f"Error sending email: {str(e)}"

#---------------------------
# Main Application
#---------------------------
def main():
    if not check_password():
        st.stop()

    # --- Branding Header with Logo ---
    st.markdown("""
    <div class="branding-header">
        <div style="display:flex; align-items:center; gap:18px;">
            <svg width="70" height="70" viewBox="0 0 70 70" xmlns="http://www.w3.org/2000/svg">
                <rect width="70" height="70" rx="14" fill="white"/>
                <text x="35" y="40" text-anchor="middle" font-family="Arial Black, Arial" font-size="28" font-weight="900" fill="#1B2A4A">RM</text>
                <polyline points="8,58 20,48 32,53 48,38 62,28" stroke="#5CACEE" stroke-width="3.5" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
                <circle cx="62" cy="28" r="3.5" fill="#5CACEE"/>
            </svg>
            <div>
                <h1 style="margin:0; font-size:28px; color:white;">RealtyMetric Solutions</h1>
                <p style="margin:4px 0 0 0; color:#A8C4E0; font-size:14px;">Agent Recruiting Dashboard – California</p>
            </div>
        </div>
        <div style="text-align:right; color:#A8C4E0; font-size:12px;">
            <p style="margin:0;">Version 1.4</p>
            <p style="margin:2px 0 0 0;">© 2026 RealtyMetric Solutions</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    This interactive tool helps brokerage teams identify promising real estate agents for recruitment.
    Rows with missing fields are kept and scored conservatively so you see more people.
    """)

    # --- About This Tool ---
    with st.expander("📖 About This Tool"):
        st.markdown("""
        ### What Is This Tool?
        RealtyMetric Solutions' Agent Recruiting Dashboard is a data-driven platform designed to help 
        real estate brokerages identify, evaluate, and recruit top-performing agents. It analyzes publicly 
        available agent activity data and generates an objective recruiting score for each agent, helping 
        your team prioritize outreach efforts.

        ---

        ### How Does the Scoring Work?
        Each agent is scored on a **0 to 1 scale** based on four weighted components:

        | Factor | Description | What It Means |
        |---|---|---|
        | **Volume** | Agent's total transaction volume | Higher volume = stronger producer |
        | **DOM (Days on Market)** | Average days properties stay listed | Lower DOM = more efficient agent |
        | **Recency** | How recently the agent completed a transaction | More recent = more active agent |
        | **Office Changes** | Whether the agent has switched offices | Recent switchers may be open to moving again |

        ---

        ### How Are Weights Used?
        The sliders in the sidebar control how much each factor contributes to the **Final Score**. 
        For example, if your brokerage values high volume producers, increase the Volume weight. 
        If you prefer agents who close quickly, increase the DOM weight. The weights do **not** need 
        to add up to 1.0 — the system will calculate the score based on whatever combination you set.

        ---

        ### How Should I Interpret the Scores?
        - **0.8 – 1.0** → Highly promising recruits — prioritize outreach
        - **0.6 – 0.8** → Strong candidates — worth investigating further
        - **0.4 – 0.6** → Moderate candidates — consider for long-term pipeline
        - **Below 0.4** → Lower priority at this time

        ---

        ### What About Missing Data?
        Agents with missing fields are **not removed** from the results. Instead, missing values are 
        scored conservatively (treated as the least favorable value), so these agents still appear but 
        rank lower. This ensures you don't accidentally overlook promising agents due to incomplete records.
        """)

    # --- How to Use ---
    with st.expander("🛈 How to Use This Tool"):
        st.markdown("""
        - **Quick Filters:** Use one-click preset buttons for common searches:
          - 🏆 Top 10%: Shows top 10% of agents by score
          - 🔄 Recent Movers: Agents who moved within 90 days
          - 💰 High Volume: Agents above median volume
          - ⚡ Low DOM: Agents below median days on market
        - **Clear All Filters:** Reset all filters and weights to defaults
        - **Adjust Weights:** Use the sidebar sliders to set how much each factor matters to your recruiting criteria.
        - **Filter by County/Office:** Select specific markets or brokerages to focus your search.
        - **Advanced Filters Toggle:** Show or hide detailed threshold filters.
        - **Set Thresholds:** Use Min Volume, Max DOM, and Max Recency to exclude agents who don't meet your minimums.
        - **Search Agents:** Use the search bar to quickly find a specific agent by name.
        - **Read the Charts:** The Top 10 chart shows your best candidates at a glance. The scatter plot shows score vs. volume.
        - **Export Data:** Download results as CSV or Excel for further analysis or sharing with your team.
        - **Missing Values:** Filters allow missing values — agents with missing DOM or Recency are kept but scored conservatively.
        """)

    # --- Disclaimer / Terms of Use ---
    with st.expander("⚖️ Disclaimer & Terms of Use"):
        st.markdown("""
        ### Disclaimer
        RealtyMetric Solutions provides this dashboard for **informational purposes only**. The data and 
        scores presented are based on publicly available information and proprietary algorithms, but are 
        not guaranteed to be accurate, complete, or up to date. RealtyMetric Solutions makes no warranties, 
        express or implied, regarding the quality or fitness of this tool for any particular purpose.

        ---

        ### Terms of Use
        By using this dashboard, you agree to the following:

        1. **Confidentiality:** All data displayed in this dashboard is confidential. You may not share, 
           distribute, or publish any information obtained from this tool without the prior written consent 
           of RealtyMetric Solutions.

        2. **No Redistribution:** You may not copy, reproduce, or distribute the content of this dashboard 
           or any exports without authorization.

        3. **Authorized Use Only:** This tool is intended solely for the authorized user or organization. 
           Sharing login credentials or access with unauthorized parties is strictly prohibited.

        4. **No Guarantee of Results:** Recruiting outcomes based on this tool are not guaranteed. 
           RealtyMetric Solutions is not liable for any business decisions made based on the data or 
           scores provided.

        5. **Data Accuracy:** While we strive to provide accurate data, RealtyMetric Solutions is not 
           responsible for any errors or omissions in the data. Users should independently verify 
           information before making critical business decisions.

        6. **Intellectual Property:** The scoring methodology, algorithms, and dashboard design are the 
           intellectual property of RealtyMetric Solutions and are protected by law.

        7. **Changes:** RealtyMetric Solutions reserves the right to update, modify, or discontinue 
           this service at any time without prior notice.

        ---

        *By accessing this dashboard, you acknowledge and agree to these terms. For questions regarding 
        these terms, please contact support@realtymetricsolutions.com*
        """)

    # Load data
    agents, city2county_full = load_and_prepare_data()
    if agents is None:
        st.stop()

    # --- Sidebar ---
    with st.sidebar:
        st.header("⚡ Quick Filters")
        st.caption("One-click presets for common searches")
        
        col_q1, col_q2 = st.columns(2)
        with col_q1:
            top_performers = st.button("🏆 Top 10%", use_container_width=True, help="Top performers by score")
            high_volume = st.button("💰 High Volume", use_container_width=True, help="Above median volume")
        with col_q2:
            recent_movers = st.button("🔄 Recent Movers", use_container_width=True, help="Moved within 90 days")
            low_dom = st.button("⚡ Low DOM", use_container_width=True, help="Below median DOM")
        
        # Clear filters button
        if st.button("🔃 Clear All Filters", use_container_width=True):
            st.rerun()
        
        st.divider()
        
        st.header("⚖️ Scoring Weights")
        volume_weight   = st.slider("Weight: Volume", 0.0, 1.0, 0.30, 0.05)
        dom_weight      = st.slider("Weight: DOM (Lower is Better)", 0.0, 1.0, 0.35, 0.05)
        recency_weight  = st.slider("Weight: Recency (Recent Better)", 0.0, 1.0, 0.20, 0.05)
        change_weight   = st.slider("Weight: Office Changes", 0.0, 1.0, 0.20, 0.05)
        
        st.divider()
        
        st.header("🔍 Filters")
        
        # Agent Search
        search_query = st.text_input("🔎 Search by agent name...", placeholder="e.g. John Smith")
        
        # Advanced filters toggle (default visible to avoid session state conflicts)
        show_advanced = st.checkbox("Show Advanced Filters", value=True)
        
        if show_advanced:
            st.caption("Rows with missing fields are kept")
            
            # County filter
            counties = sorted(agents['County'].dropna().unique())
            county_filter = st.multiselect(
                "County (optional):", 
                options=counties, 
                default=None,
                help="Leave empty to include all counties"
            )
            
            # Office filter
            offices = sorted(agents['OfficeName'].dropna().unique())
            offices = [o for o in offices if str(o).strip() != '']
            office_filter = st.multiselect(
                "Office (optional):",
                options=offices,
                default=None,
                help="Leave empty to include all offices"
            )
            
            min_volume  = st.number_input("Min Total Volume ($):", min_value=0, value=100000, step=100000, format="%d")
            max_dom     = st.number_input("Max Days on Market:", min_value=1, value=180, step=10)
            max_recency = st.number_input("Max Days Since Move:", min_value=1, value=999, step=50)
        else:
            # Use defaults when advanced filters hidden
            county_filter = []
            office_filter = []
            min_volume = 100000
            max_dom = 180
            max_recency = 999

    # --- Process Data ---
    def process_agents():
        df0 = agents.copy()
        total_rows = len(df0)
        
        # County filter
        df1 = df0[df0['County'].isin(county_filter)] if county_filter else df0.copy()
        
        # Office filter
        if office_filter:
            df1 = df1[df1['OfficeName'].isin(office_filter)]
        
        # Apply threshold filters (keep rows with NAs)
        df2 = df1.copy()
        if 'TotalVolume'  in df2.columns: df2 = df2[(df2['TotalVolume'].isna())  | (df2['TotalVolume']  >= min_volume)]
        if 'DaysOnMarket' in df2.columns: df2 = df2[(df2['DaysOnMarket'].isna()) | (df2['DaysOnMarket'] <= max_dom)]
        if 'RecencyDays'  in df2.columns: df2 = df2[(df2['RecencyDays'].isna())  | (df2['RecencyDays']  <= max_recency)]

        df3 = df2.copy()
        df3['vol_for_score']     = df3['TotalVolume'].fillna(0) if 'TotalVolume' in df3.columns else 0
        df3['dom_for_score']     = df3['DaysOnMarket'].fillna(max_dom) if 'DaysOnMarket' in df3.columns else max_dom
        df3['recency_for_score'] = df3['RecencyDays'].fillna(max_recency) if 'RecencyDays' in df3.columns else max_recency

        df3['Volume_Score']  = safe_rescale(df3['vol_for_score'])
        df3['DOM_Score']     = safe_rescale(-df3['dom_for_score'])
        df3['Recency_Score'] = safe_rescale(-df3['recency_for_score'])

        if 'PreviousOfficeName' in df3.columns:
            has_prev = (~df3['PreviousOfficeName'].isna()) & (df3['PreviousOfficeName'] != '')
            df3['Change_Score'] = safe_rescale(has_prev.astype(int))
        else:
            df3['Change_Score'] = 0

        df3['Final_Score'] = (
            volume_weight  * df3['Volume_Score']  +
            dom_weight     * df3['DOM_Score']     +
            recency_weight * df3['Recency_Score'] +
            change_weight  * df3['Change_Score']
        )

        df_final = df3.sort_values('Final_Score', ascending=False).reset_index(drop=True)
        df_final['Rank'] = df_final.index + 1

        cols = ['Rank','FirstName','LastName','OfficeName','County','TotalVolume','DaysOnMarket','RecencyDays','Final_Score']
        cols = [c for c in cols if c in df_final.columns]
        df_display = df_final[cols].copy()
        
        # Apply Quick Filters
        if top_performers:
            # Top 10% by score
            if len(df_display) > 0:
                threshold = df_display['Final_Score'].quantile(0.90)
                df_display = df_display[df_display['Final_Score'] >= threshold].reset_index(drop=True)
                df_display['Rank'] = df_display.index + 1
                st.info(f"🏆 Showing Top 10% (Score ≥ {threshold:.4f})")
        
        if recent_movers:
            # Moved within 90 days
            if 'RecencyDays' in df_display.columns:
                df_display = df_display[(df_display['RecencyDays'].isna()) | (df_display['RecencyDays'] <= 90)].reset_index(drop=True)
                df_display['Rank'] = df_display.index + 1
                st.info("🔄 Showing Recent Movers (moved within 90 days)")
        
        if high_volume:
            # Above median volume
            if 'TotalVolume' in df_display.columns:
                median_vol = df_display['TotalVolume'].median()
                df_display = df_display[(df_display['TotalVolume'].isna()) | (df_display['TotalVolume'] >= median_vol)].reset_index(drop=True)
                df_display['Rank'] = df_display.index + 1
                st.info(f"💰 Showing High Volume Agents (≥ ${median_vol:,.0f})")
        
        if low_dom:
            # Below median DOM
            if 'DaysOnMarket' in df_display.columns:
                median_dom = df_display['DaysOnMarket'].median()
                df_display = df_display[(df_display['DaysOnMarket'].isna()) | (df_display['DaysOnMarket'] <= median_dom)].reset_index(drop=True)
                df_display['Rank'] = df_display.index + 1
                st.info(f"⚡ Showing Low DOM Agents (≤ {int(median_dom)} days)")

        # Agent search filter
        if search_query and search_query.strip():
            q = search_query.strip().lower()
            mask = (
                df_display['FirstName'].str.lower().str.contains(q, na=False) |
                df_display['LastName'].str.lower().str.contains(q, na=False) |
                (df_display['FirstName'].str.lower() + ' ' + df_display['LastName'].str.lower()).str.contains(q, na=False)
            )
            df_display = df_display[mask].reset_index(drop=True)
            df_display['Rank'] = df_display.index + 1

        return df_display, df_final, total_rows, len(df1), len(df2)

    df_display, df_full, total_rows, after_county, after_filters = process_agents()

    # --- Summary Stats Cards ---
    st.subheader("📊 Summary Overview")
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.metric("Total Agents", f"{total_rows:,}")
    with c2:
        st.metric("After Filters", f"{len(df_display):,}")
    with c3:
        avg_vol = df_display['TotalVolume'].mean() if 'TotalVolume' in df_display.columns else 0
        st.metric("Avg Volume", f"${avg_vol:,.0f}" if not pd.isna(avg_vol) else "—")
    with c4:
        med_dom = df_display['DaysOnMarket'].median() if 'DaysOnMarket' in df_display.columns else 0
        st.metric("Median DOM", f"{int(med_dom)}" if not pd.isna(med_dom) else "—")
    with c5:
        top_score = df_display['Final_Score'].max() if not df_display.empty else 0
        st.metric("Top Score", f"{top_score:.4f}" if not pd.isna(top_score) else "—")

    if len(df_display) == 0:
        st.warning("No agents match the current filters.")
        return

    # --- Row Audit ---
    st.subheader("🔍 Row Audit")
    a1, a2, a3, a4 = st.columns(4)
    with a1: st.metric("Rows in CSV", f"{total_rows:,}")
    with a2: 
        label = "After County/Office" if (county_filter or office_filter) else "After Filters"
        st.metric(label, f"{after_county:,}")
    with a3: st.metric("After Thresholds", f"{after_filters:,}")
    with a4: st.metric("Rows Shown", f"{len(df_display):,}")

    # --- Scatter Plot (colored by Score) ---
    st.subheader("📈 Score vs. Total Volume")
    if 'TotalVolume' in df_display.columns:
        scatter_df = df_display.dropna(subset=['TotalVolume']).copy()

        fig_scatter = px.scatter(
            scatter_df, x='TotalVolume', y='Final_Score', color='Final_Score',
            color_continuous_scale='Viridis',
            hover_data={
                'Rank': True, 'FirstName': True, 'LastName': True,
                'County': True, 'OfficeName': True,
                'TotalVolume': ':$,.0f', 'DaysOnMarket': ':.0f',
                'RecencyDays': ':.0f', 'Final_Score': ':.3f'
            },
            title="Recruiting Score vs Total Volume"
        )

        fig_scatter.update_layout(
            xaxis_title="Total Volume ($)", yaxis_title="Recruiting Score",
            height=500, xaxis=dict(tickprefix="$", tickformat=",")
        )
        fig_scatter.update_traces(marker=dict(size=9))
        st.plotly_chart(fig_scatter, use_container_width=True)

    # --- Ranked Agent Table ---
    st.subheader("📋 Ranked Agent Table")
    df_formatted = df_display.copy()
    if 'TotalVolume'  in df_formatted.columns: df_formatted['TotalVolume']  = df_formatted['TotalVolume'].apply(format_currency)
    if 'DaysOnMarket' in df_formatted.columns: df_formatted['DaysOnMarket'] = df_formatted['DaysOnMarket'].apply(format_integer)
    if 'RecencyDays'  in df_formatted.columns: df_formatted['RecencyDays']  = df_formatted['RecencyDays'].apply(format_integer)
    if 'Final_Score'  in df_formatted.columns: df_formatted['Final_Score']  = df_formatted['Final_Score'].round(4)

    st.dataframe(
        df_formatted, use_container_width=True, hide_index=True,
        column_config={
            "Rank":          st.column_config.NumberColumn("Rank", format="%d"),
            "FirstName":     "First Name",
            "LastName":      "Last Name",
            "OfficeName":    "Office Name",
            "County":        "County",
            "TotalVolume":   "Total Volume",
            "DaysOnMarket":  "Days on Market",
            "RecencyDays":   "Days Since Move",
            "Final_Score":   st.column_config.NumberColumn("Final Score", format="%.4f")
        }
    )

    # --- Export Buttons ---
    st.subheader("📥 Export & Reporting")
    
    # Column selector
    with st.expander("⚙️ Customize Export Columns"):
        st.caption("Select which columns to include in your export")
        available_cols = df_display.columns.tolist()
        default_cols = ['Rank', 'FirstName', 'LastName', 'OfficeName', 'County', 'TotalVolume', 'Final_Score']
        default_cols = [c for c in default_cols if c in available_cols]
        selected_cols = st.multiselect(
            "Columns to export:",
            options=available_cols,
            default=default_cols,
            help="Choose which columns to include in your exported reports"
        )
        if not selected_cols:
            st.warning("Please select at least one column to export")
            selected_cols = default_cols
    
    export_df = df_display[selected_cols].copy()
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
    
    # Export buttons
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**Standard Exports**")
        csv_buffer = StringIO()
        export_df.to_csv(csv_buffer, index=False)
        st.download_button(
            label="📄 Download CSV",
            data=csv_buffer.getvalue(),
            file_name=f"RealtyMetric_CA_Agents_{timestamp}.csv",
            mime="text/csv",
            use_container_width=True
        )
        
        xlsx_buffer = export_to_excel(export_df)
        st.download_button(
            label="📊 Download Excel",
            data=xlsx_buffer,
            file_name=f"RealtyMetric_CA_Agents_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        pdf_buffer = export_to_pdf(export_df)
        st.download_button(
            label="📑 Download PDF",
            data=pdf_buffer,
            file_name=f"RealtyMetric_CA_Agents_{timestamp}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    
    with col2:
        st.markdown("**Executive Summary**")
        st.caption("Top 20 agents")
        
        exec_excel = export_to_excel(export_df, include_summary=True, top_n=20)
        st.download_button(
            label="📊 Summary (Excel)",
            data=exec_excel,
            file_name=f"RealtyMetric_Executive_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        exec_pdf = export_to_pdf(export_df, charts_data=None, top_n=20)
        st.download_button(
            label="📑 Summary (PDF)",
            data=exec_pdf,
            file_name=f"RealtyMetric_Executive_{timestamp}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    
    with col3:
        user_tier = st.secrets.get("tier", "starter")
        
        if user_tier in ["professional", "enterprise"]:
            st.markdown("**📧 Email Report** 🌟")
            st.caption("Professional feature")
            
            recipient_email = st.text_input(
                "Recipient email:",
                placeholder="client@example.com"
            )
            
            report_type = st.radio(
                "Report type:",
                ["Standard Excel", "Executive Summary"]
            )
            
            if st.button("📧 Send Email", use_container_width=True):
                if not recipient_email or '@' not in recipient_email:
                    st.error("Please enter a valid email address")
                else:
                    with st.spinner("Sending email..."):
                        if report_type == "Standard Excel":
                            attachment = export_to_excel(export_df).read()
                            filename = f"RealtyMetric_CA_{timestamp}.xlsx"
                        else:
                            attachment = export_to_excel(export_df, include_summary=True, top_n=20).read()
                            filename = f"RealtyMetric_Executive_{timestamp}.xlsx"
                        
                        success, message = send_email_report(recipient_email, attachment, filename)
                        
                        if success:
                            st.success(f"✅ {message}")
                        else:
                            st.error(f"❌ {message}")
        else:
            st.markdown("**📧 Email Reports** 🔒")
            st.caption("Upgrade to Professional")
            st.info("Email reporting is available in Professional tier. Contact sales@realtymetricsolutions.com")
    
    st.divider()

    # --- Footer ---
    st.markdown("""
    <div class="footer">
        <p>© 2026 RealtyMetric Solutions | Agent Recruiting Dashboard | Confidential</p>
        <p>For support, contact: support@realtymetricsolutions.com | Version 1.4</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
